import re
import csv
import random
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import load_workbook

BASE_DIR = Path(__file__).parent
ROSTERS_DIR = BASE_DIR / "rosters"
PREVIOUS_DIR = BASE_DIR / "previous_schedules"
CONFIG_FILE = BASE_DIR / "config.txt"
GROUPS_FILE = BASE_DIR / "groups.txt"
UNAVAIL_FILE = BASE_DIR / "unavailability.txt"
OVERRIDES_FILE = BASE_DIR / "overrides.txt"


def parse_config():
    roles = {}
    start = end = None
    seed = 42
    with open(CONFIG_FILE) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if line.startswith("scheduling_start:"):
                start = datetime.strptime(line.split(":", 1)[1].strip(), "%Y-%m-%d").date()
            elif line.startswith("scheduling_end:"):
                end = datetime.strptime(line.split(":", 1)[1].strip(), "%Y-%m-%d").date()
            elif line.startswith("seed:"):
                seed = int(line.split(":", 1)[1].strip())
            else:
                match = re.match(
                    r"([\w-]+):\s*min_gap=(\d+),\s*type=(individual|group),\s*week=(odd|even|all)"
                    r"(?:,\s*service_type=(\w+))?(?:,\s*hidden=(true|false))?",
                    line,
                )
                if match:
                    roles[match.group(1)] = {
                        "min_gap": int(match.group(2)),
                        "type": match.group(3),
                        "week": match.group(4),
                        "service_type": match.group(5),
                        "hidden": match.group(6) == "true",
                    }
    return roles, start, end, seed


def parse_groups():
    groups = {}
    if not GROUPS_FILE.exists():
        return groups
    with open(GROUPS_FILE) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            name, members = line.split(":", 1)
            groups[name.strip()] = [m.strip() for m in members.split(",")]
    return groups


def parse_roster(role_name):
    path = ROSTERS_DIR / f"{role_name}.txt"
    candidates = []
    if not path.exists():
        return candidates
    with open(path) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            candidates.append(line)
    return candidates


def parse_unavailability():
    unavail = {}
    if not UNAVAIL_FILE.exists():
        return unavail
    with open(UNAVAIL_FILE) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            name, dates_str = line.split(":", 1)
            name = name.strip()
            if name not in unavail:
                unavail[name] = set()
            for part in dates_str.split(","):
                part = part.strip()
                if " to " in part:
                    s, e = part.split(" to ")
                    sd = datetime.strptime(s.strip(), "%Y-%m-%d").date()
                    ed = datetime.strptime(e.strip(), "%Y-%m-%d").date()
                    d = sd
                    while d <= ed:
                        unavail[name].add(d)
                        d += timedelta(days=1)
                else:
                    unavail[name].add(datetime.strptime(part, "%Y-%m-%d").date())
    return unavail


def parse_overrides(roles):
    """Returns {(date, role): candidate} for one-off locked-in assignments.

    Raises ValueError on malformed lines, unknown roles, non-Saturday dates,
    or duplicate overrides for the same role/date.
    """
    overrides = {}
    if not OVERRIDES_FILE.exists():
        return overrides
    with open(OVERRIDES_FILE) as f:
        for lineno, line in enumerate(f, start=1):
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = [p.strip() for p in line.split(":")]
            if len(parts) != 3 or not all(parts):
                raise ValueError(
                    f"overrides.txt line {lineno}: expected 'role: YYYY-MM-DD: name', got: {line}"
                )
            role, date_str, name = parts
            if role not in roles:
                raise ValueError(f"overrides.txt line {lineno}: unknown role '{role}'")
            try:
                date = datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                raise ValueError(f"overrides.txt line {lineno}: invalid date '{date_str}'")
            if date.weekday() != 5:
                raise ValueError(f"overrides.txt line {lineno}: {date_str} is not a Saturday")
            key = (date, role)
            if key in overrides:
                raise ValueError(
                    f"overrides.txt line {lineno}: duplicate override for role '{role}' on {date_str}"
                )
            overrides[key] = name
    return overrides


def load_previous_schedules(roles):
    """Returns {role: {candidate: [list of dates served]}}"""
    history = {role: {} for role in roles}
    if not PREVIOUS_DIR.exists():
        return history

    def process_rows(headers, rows):
        date_col = 0
        role_cols = {}
        for i, h in enumerate(headers):
            if h and h.strip().lower() == "date":
                date_col = i
            elif h and h.strip().lower().replace(" ", "-") in roles:
                role_cols[h.strip().lower().replace(" ", "-")] = i
        for row in rows:
            date_val = row[date_col]
            if isinstance(date_val, datetime):
                date_val = date_val.date()
            elif isinstance(date_val, str) and date_val.strip():
                try:
                    date_val = datetime.strptime(date_val.strip(), "%Y-%m-%d").date()
                except ValueError:
                    continue
            else:
                continue
            for role, col in role_cols.items():
                if col < len(row) and row[col]:
                    candidate = str(row[col]).strip()
                    if candidate:
                        if candidate not in history[role]:
                            history[role][candidate] = []
                        history[role][candidate].append(date_val)

    for f in sorted(PREVIOUS_DIR.iterdir()):
        if f.suffix == ".csv":
            with open(f, newline="", encoding="utf-8") as csvfile:
                reader = csv.reader(csvfile)
                headers = next(reader)
                process_rows(headers, list(reader))
        elif f.suffix == ".xlsx":
            wb = load_workbook(f, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            process_rows(headers, list(ws.iter_rows(min_row=2, values_only=True)))
            wb.close()

    return history


def get_saturdays(start, end):
    dates = []
    d = start
    while d.weekday() != 5:
        d += timedelta(days=1)
    while d <= end:
        dates.append(d)
        d += timedelta(days=7)
    return dates


def saturday_week_number(date):
    """Returns which Saturday of the month this is (1-based)."""
    count = 0
    d = date.replace(day=1)
    while d <= date:
        if d.weekday() == 5:
            count += 1
        d += timedelta(days=1)
    return count


def is_active_week(date, week_type):
    """Check if a role is active on this Saturday based on week type."""
    if week_type == "all":
        return True
    n = saturday_week_number(date)
    if week_type == "odd":
        return n % 2 == 1
    if week_type == "even":
        return n % 2 == 0
    return True


def get_sibling_roles(role, roles):
    """Get all roles sharing the same service_type (excluding self)."""
    st = roles[role].get("service_type")
    if not st:
        return []
    return [r for r in roles if r != role and roles[r].get("service_type") == st]


def schedule(roles, start, end, groups, unavail, history, overrides=None):
    if overrides is None:
        overrides = {}
    saturdays = get_saturdays(start, end)
    saturday_set = set(saturdays)
    for (odate, orole) in overrides:
        if odate not in saturday_set:
            raise ValueError(
                f"overrides.txt: {odate} for role '{orole}' is outside the scheduling range"
            )
    schedule_result = {date: {} for date in saturdays}

    # Track last assignment date per role per candidate
    last_assigned = {}
    for role in roles:
        last_assigned[role] = {}
        if role in history:
            for cand, dates in history[role].items():
                if dates:
                    last_assigned[role][cand] = max(dates)

    # Track total assignment count per service_type for load balancing
    total_count = {}
    for role in roles:
        total_count[role] = {}
        if role in history:
            for cand, dates in history[role].items():
                total_count[role][cand] = len(dates)

    for date in saturdays:
        used_individuals = set()  # individuals used today (hard constraint)
        used_groups = set()  # group names used today (hard constraint)

        # Apply locked-in overrides for this date first, before the algorithm runs.
        date_overrides = {role: cand for (d, role), cand in overrides.items() if d == date}
        for role, cand in date_overrides.items():
            cfg = roles[role]
            if cfg["type"] == "individual" and cand in unavail and date in unavail[cand]:
                raise ValueError(
                    f"Override conflict: {cand} is marked unavailable on {date} but pinned to {role}"
                )
            if cfg["type"] == "individual" and cand in used_individuals:
                raise ValueError(
                    f"Override conflict: {cand} is already assigned another role on {date}"
                )
            if cfg["type"] == "group" and cand in used_groups:
                raise ValueError(
                    f"Override conflict: group {cand} is already assigned another role on {date}"
                )

            schedule_result[date][role] = cand
            last_assigned[role][cand] = date
            total_count[role][cand] = total_count[role].get(cand, 0) + 1

            if cfg["type"] == "individual":
                used_individuals.add(cand)
            elif cfg["type"] == "group":
                used_groups.add(cand)
                if cand in groups:
                    for m in groups[cand]:
                        used_individuals.add(m)

        # Sort roles by number of eligible candidates (most constrained first)
        # Hidden roles (e.g. pembawa-khotbah) are never auto-assigned; they only
        # get a value through overrides.txt, and only feed sibling gap/count tracking.
        active_roles = [
            r for r in roles
            if is_active_week(date, roles[r]["week"]) and r not in date_overrides and not roles[r]["hidden"]
        ]
        role_order = sorted(active_roles, key=lambda r: len(parse_roster(r)))

        for role in role_order:
            cfg = roles[role]
            candidates = parse_roster(role)
            min_gap_days = cfg["min_gap"] * 7
            siblings = get_sibling_roles(role, roles)
            eligible = []

            for cand in candidates:
                # Unavailability only blocks individual assignments, not groups
                if cfg["type"] == "individual" and cand in unavail and date in unavail[cand]:
                    continue

                # Hard: individual not already used today
                if cfg["type"] == "individual" and cand in used_individuals:
                    continue

                # Hard: group not already used today in another role
                if cfg["type"] == "group" and cand in used_groups:
                    continue

                # Check gap for this role
                if cand in last_assigned[role]:
                    days_since = (date - last_assigned[role][cand]).days
                    if days_since < min_gap_days:
                        continue

                # Check gap across sibling roles (shared service_type)
                gap_violated = False
                for sib in siblings:
                    if cand in last_assigned.get(sib, {}):
                        days_since = (date - last_assigned[sib][cand]).days
                        if days_since < min_gap_days:
                            gap_violated = True
                            break
                if gap_violated:
                    continue

                eligible.append(cand)

            if not eligible:
                schedule_result[date][role] = ""
                continue

            # Score: combine counts across sibling roles for load balancing
            def score(cand, _role=role, _siblings=siblings):
                count = total_count[_role].get(cand, 0)
                for sib in _siblings:
                    count += total_count.get(sib, {}).get(cand, 0)
                # Longest since last across role + siblings
                last = None
                if cand in last_assigned[_role]:
                    last = last_assigned[_role][cand]
                for sib in _siblings:
                    if cand in last_assigned.get(sib, {}):
                        sib_last = last_assigned[sib][cand]
                        if last is None or sib_last > last:
                            last = sib_last
                days_since = (date - last).days if last else 9999
                return (count, -days_since, random.random())

            eligible.sort(key=score)
            chosen = eligible[0]
            schedule_result[date][role] = chosen
            last_assigned[role][chosen] = date
            total_count[role][chosen] = total_count[role].get(chosen, 0) + 1

            # Mark used
            if cfg["type"] == "individual":
                used_individuals.add(chosen)
            elif cfg["type"] == "group":
                used_groups.add(chosen)
                # Soft: mark group members as used individually
                if chosen in groups:
                    for m in groups[chosen]:
                        used_individuals.add(m)

    return schedule_result, saturdays


def write_output(schedule_result, saturdays, roles, overrides=None):
    import csv

    if overrides is None:
        overrides = {}
    # Hidden roles don't get a column - they only exist to feed sibling gap/count
    # tracking for another service_type when pinned via overrides.txt.
    role_list = [r for r in roles if not roles[r]["hidden"]]
    headers = ["Date"] + role_list

    csv_path = BASE_DIR / "schedule_output.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for date in saturdays:
            row = [date.strftime("%Y-%m-%d")]
            for role in role_list:
                val = schedule_result[date].get(role, "")
                if not is_active_week(date, roles[role]["week"]) and (date, role) not in overrides:
                    val = ""
                row.append(val if val else "")
            writer.writerow(row)
    print(f"Schedule saved to: {csv_path}")


def main():
    roles, start, end, seed = parse_config()
    random.seed(seed)
    groups = parse_groups()
    unavail = parse_unavailability()
    overrides = parse_overrides(roles)
    history = load_previous_schedules(roles)
    result, saturdays = schedule(roles, start, end, groups, unavail, history, overrides)
    write_output(result, saturdays, roles, overrides)


if __name__ == "__main__":
    main()
